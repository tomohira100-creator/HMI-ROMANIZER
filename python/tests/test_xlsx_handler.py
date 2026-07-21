"""XLSX handler integration tests.

Same three-tier strategy as the DOCX handler:
  (a) untouched parts byte-identical
  (b) edited parts structurally isomorphic, text aside
  (c) the text is correct

The byte-identity tier is meaningful for XLSX only because shared strings are
romanized in place: cells reference them by index, and the index is preserved,
so no worksheet moves on account of the strings.
"""

import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from lxml import etree

from romanizer.handlers import xlsx_handler as X

M = X.M

REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLES = REPO_ROOT / "samples"
BUILDER = REPO_ROOT / "python" / "tests" / "fixtures" / "build_fixtures.py"
FIXTURE = SAMPLES / "12_spreadsheet.xlsx"

REAL = SAMPLES / "メゾンドオーラ中目黒２F見積書20260605.xlsx"


@pytest.fixture(scope="module", autouse=True)
def ensure_fixture():
    if not FIXTURE.exists():
        subprocess.run([sys.executable, str(BUILDER)], check=True)


def parts(path):
    with zipfile.ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def skeleton(blob):
    """(tag, sorted-attrs) per element, ignoring text. Structure must survive."""
    out = []
    for element in etree.fromstring(blob).iter():
        out.append((element.tag, tuple(sorted(element.attrib.items()))))
    return out


def strings(blob):
    root = etree.fromstring(blob)
    return ["".join(t.text or "" for t in si.iter(M + "t")) for si in root.iter(M + "si")]


# --- Synthetic fixture -----------------------------------------------------

@pytest.fixture(scope="module")
def converted(tmp_path_factory):
    out = tmp_path_factory.mktemp("xlsx") / "out.xlsx"
    X.convert(FIXTURE, out)
    return out


def test_no_parts_added_or_lost(converted):
    assert set(parts(FIXTURE)) == set(parts(converted))


def test_untouched_parts_are_byte_identical(converted):
    before, after = parts(FIXTURE), parts(converted)
    for name in ("xl/media/image1.png", "xl/styles.xml"):
        assert before[name] == after[name], name


def test_only_expected_parts_changed(converted):
    before, after = parts(FIXTURE), parts(converted)
    changed = {n for n in before if before[n] != after[n]}
    assert changed == {
        "xl/sharedStrings.xml", "xl/workbook.xml",
        "xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml",
    }


def test_shared_strings_romanized_in_place(converted):
    # Same count, same order -- indices preserved.
    assert strings(parts(converted)["xl/sharedStrings.xml"]) == [
        "Kabushiki Gaisha", "Tōkyō", "Gōkei",
    ]


def test_shared_string_count_and_order_unchanged(converted):
    before = etree.fromstring(parts(FIXTURE)["xl/sharedStrings.xml"])
    after = etree.fromstring(parts(converted)["xl/sharedStrings.xml"])
    assert len(before.findall(M + "si")) == len(after.findall(M + "si"))


def test_sheet_XML_structure_is_isomorphic(converted):
    for name in ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"):
        assert skeleton(parts(FIXTURE)[name]) == skeleton(parts(converted)[name]), name


def test_sheet_names_romanized(converted):
    wb = etree.fromstring(parts(converted)["xl/workbook.xml"])
    assert [s.get("name") for s in wb.iter(M + "sheet")] == ["Uchiwake", "Hyōshi"]


def test_defined_names_references_rewritten_and_quoted(converted):
    wb = etree.fromstring(parts(converted)["xl/workbook.xml"])
    refs = [d.text for d in wb.iter(M + "definedName")]
    assert "'Hyōshi'!$A$1:$B$1" in refs      # was '表紙 '! (quoted, trailing space)
    assert "'Uchiwake'!$A$1:$E$1" in refs    # was 内訳! (unquoted)


def test_cross_sheet_formula_references_rewritten(converted):
    s1 = etree.fromstring(parts(converted)["xl/worksheets/sheet1.xml"])
    s2 = etree.fromstring(parts(converted)["xl/worksheets/sheet2.xml"])
    f1 = {c.get("r"): c.find(M + "f").text for c in s1.iter(M + "c") if c.find(M + "f") is not None}
    f2 = {c.get("r"): c.find(M + "f").text for c in s2.iter(M + "c") if c.find(M + "f") is not None}
    assert f1["B1"] == "'Hyōshi'!A1"     # quoted cross-ref
    assert f2["B1"] == "'Uchiwake'!A1"   # unquoted -> quoted
    assert f1["C1"] == 'A1&"Gōkei"'      # formula literal romanized


def test_cached_string_value_romanized(converted):
    s1 = etree.fromstring(parts(converted)["xl/worksheets/sheet1.xml"])
    d1 = next(c for c in s1.iter(M + "c") if c.get("r") == "D1")
    assert d1.get("t") == "str"
    assert d1.find(M + "v").text == "Tōkyō"


def test_no_japanese_remains_anywhere(converted):
    for name, blob in parts(converted).items():
        if name.endswith(".xml"):
            assert not X._JAPANESE.search(blob.decode("utf-8", "replace")), name


def test_output_reopens(converted):
    import openpyxl

    wb = openpyxl.load_workbook(converted)
    assert wb.sheetnames == ["Uchiwake", "Hyōshi"]


def test_idempotent(tmp_path):
    once, twice = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    X.convert(FIXTURE, once)
    X.convert(once, twice)
    assert strings(parts(once)["xl/sharedStrings.xml"]) == strings(parts(twice)["xl/sharedStrings.xml"])


def test_custom_dictionary_is_honoured(tmp_path):
    import json
    from romanizer import dictionary

    (tmp_path / "custom_terms.json").write_text(
        json.dumps({"東京": "TOKYO"}, ensure_ascii=False), encoding="utf-8"
    )
    (tmp_path / "abbreviations.json").write_text('{"abbreviations": []}', encoding="utf-8")
    dic = dictionary.load(tmp_path)
    out = tmp_path / "out.xlsx"
    X.convert(FIXTURE, out, dic)
    assert "TOKYO" in strings(parts(out)["xl/sharedStrings.xml"])


# --- Real Maison d'Aura workbook (gitignored; skipped on a clean checkout) --

def test_real_workbook_preserves_untouched_parts(tmp_path):
    if not REAL.exists():
        pytest.skip("real sample not present")
    out = tmp_path / "out.xlsx"
    X.convert(REAL, out)
    before, after = parts(REAL), parts(out)
    assert set(before) == set(after)
    # Everything except shared strings, workbook, and worksheets must be
    # byte-identical -- the image, both drawings, all printer settings, styles,
    # theme, and calcChain the human's own output also kept.
    for name in before:
        edited = name == "xl/sharedStrings.xml" or name == "xl/workbook.xml" or X._SHEET.match(name)
        if not edited:
            assert before[name] == after[name], "re-serialized: {}".format(name)


def test_real_workbook_leaves_no_japanese_in_formulas(tmp_path):
    if not REAL.exists():
        pytest.skip("real sample not present")
    out = tmp_path / "out.xlsx"
    X.convert(REAL, out)
    after = parts(out)
    for name, blob in after.items():
        if X._SHEET.match(name):
            root = etree.fromstring(blob)
            for f in root.iter(M + "f"):
                if f.text:
                    assert not X._JAPANESE.search(f.text), (name, f.text)


def test_real_workbook_reopens(tmp_path):
    if not REAL.exists():
        pytest.skip("real sample not present")
    import openpyxl

    out = tmp_path / "out.xlsx"
    X.convert(REAL, out)
    openpyxl.load_workbook(out)


# --- CLI --------------------------------------------------------------------

def test_cli_converts_xlsx(tmp_path, capsys):
    from romanizer import cli

    out = tmp_path / "out.xlsx"
    assert cli.main(["convert", str(FIXTURE), str(out)]) == cli.EXIT_OK
    assert out.exists()
    assert "wrote" in capsys.readouterr().out
    assert strings(parts(out)["xl/sharedStrings.xml"]) == [
        "Kabushiki Gaisha", "Tōkyō", "Gōkei",
    ]
